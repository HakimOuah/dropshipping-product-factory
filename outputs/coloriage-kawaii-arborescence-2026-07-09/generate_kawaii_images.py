#!/usr/bin/env python3
"""Generate product carousel and storefront images from the sourcing workbook."""

from __future__ import annotations

import argparse
import base64
import csv
import os
import random
import re
import sys
import time
import unicodedata
from contextlib import ExitStack
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from openai import APIConnectionError, APITimeoutError, AuthenticationError, OpenAI, RateLimitError
from openpyxl import load_workbook
from PIL import Image


PROJECT_ROOT = Path("/Users/Hakim/Documents/New project")
WORK_DIR = PROJECT_ROOT / "outputs/coloriage-kawaii-arborescence-2026-07-09"
DEFAULT_WORKBOOK = WORK_DIR / "arborescence-site-coloriages-kawaii-sourcing-complet.xlsx"
DEFAULT_OUTPUT_ROOT = WORK_DIR / "images_generees"
PRODUCT_SHEET = "Sourcing_AliExpress"
HOME_SHEET = "Prompts_home_collection"
MODEL = "gpt-image-1"

PROMPT_LABELS = (
    "PACKSHOT",
    "LIFESTYLE",
    "KIT",
    "EN SITUATION",
    "DÉTAIL",
    "DETAIL",
    "FEATURE",
    "AVANT/APRÈS",
    "AVANT/APRES",
)
FILE_NAMES = {
    "PACKSHOT": "carousel_1_packshot.png",
    "LIFESTYLE": "carousel_2_lifestyle.png",
    "EN SITUATION": "carousel_3_en_situation.png",
    "DÉTAIL": "carousel_4_detail.png",
    "FEATURE": "carousel_5_feature.png",
}
EDIT_LABELS = {"PACKSHOT", "DÉTAIL"}
LICENSE_GUARD = (
    "Create only original generic artwork. No logos, trademarks, brand names, "
    "copyrighted characters, licensed franchises, club badges, or recognizable protected IP."
)

# Official per-image generation prices for gpt-image-1. Edit input-image tokens are extra.
COST_USD = {
    "low": {"1024x1024": 0.011, "1024x1536": 0.016, "1536x1024": 0.016},
    "medium": {"1024x1024": 0.042, "1024x1536": 0.063, "1536x1024": 0.063},
    "high": {"1024x1024": 0.167, "1024x1536": 0.250, "1536x1024": 0.250},
}


@dataclass(frozen=True)
class Product:
    row: int
    collection: str
    title: str
    verdict: str
    prompt_text: str
    reference_dir: Path


@dataclass(frozen=True)
class HomeImage:
    row: int
    zone: str
    format_label: str
    prompt: str


@dataclass(frozen=True)
class Job:
    row_label: str
    product: str
    image_type: str
    prompt: str
    size: str
    output_path: Path
    reference_paths: tuple[Path, ...] = ()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--model", default=MODEL)
    parser.add_argument("--quality", choices=("low", "medium", "high"), default="medium")
    parser.add_argument("--limit", type=int, default=0, help="Maximum sellable products; 0 means all.")
    parser.add_argument("--only-packshot", action="store_true")
    parser.add_argument("--include-home", action="store_true")
    parser.add_argument("--home-limit", type=int, default=0, help="Maximum home/collection images; 0 means all.")
    parser.add_argument("--sleep", type=float, default=1.5, help="Seconds between successful API calls.")
    parser.add_argument("--max-retries", type=int, default=5)
    parser.add_argument("--dry-run", action="store_true", help="Validate and log jobs without API calls.")
    parser.add_argument("--force", action="store_true", help="Regenerate files that already exist.")
    return parser.parse_args()


def slugify(value: str, max_length: int = 80) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")
    return (slug or "image")[:max_length].rstrip("-")


def resolve_workbook_path(value: object) -> Path:
    raw = str(value or "").strip()
    if raw.startswith(".../"):
        return PROJECT_ROOT / raw[4:]
    path = Path(raw).expanduser()
    return path if path.is_absolute() else PROJECT_ROOT / path


def normalize_label(label: str) -> str:
    upper = re.sub(r"\s+", " ", label.strip().upper())
    aliases = {
        "DETAIL": "DÉTAIL",
        "KIT": "LIFESTYLE",
        "AVANT/APRÈS": "FEATURE",
        "AVANT/APRES": "FEATURE",
    }
    return aliases.get(upper, upper)


def split_product_prompts(value: object) -> list[tuple[str, str]]:
    text = str(value or "").replace("\r\n", "\n").strip()
    label_pattern = "|".join(re.escape(label) for label in PROMPT_LABELS)
    start_re = re.compile(rf"^\s*({label_pattern})\s*(?:—|–|-|:)\s*(.*)$", re.IGNORECASE)
    prompts: list[tuple[str, str]] = []
    current_label: str | None = None
    current_lines: list[str] = []

    for line in text.splitlines():
        match = start_re.match(line)
        if match:
            if current_label:
                prompts.append((current_label, "\n".join(current_lines).strip()))
            current_label = normalize_label(match.group(1))
            current_lines = [match.group(2).strip()]
        elif current_label:
            current_lines.append(line)
    if current_label:
        prompts.append((current_label, "\n".join(current_lines).strip()))

    prompts = [(label, prompt) for label, prompt in prompts if prompt]
    expected = ["PACKSHOT", "LIFESTYLE", "EN SITUATION", "DÉTAIL", "FEATURE"]
    actual = [label for label, _ in prompts]
    if actual != expected:
        raise ValueError(f"Expected prompt labels {expected}, found {actual}")
    return prompts


def format_to_size(format_label: str) -> str:
    normalized = format_label.replace(" ", "").lower()
    if normalized in {"9:16", "4:5", "3:4", "2:3"}:
        return "1024x1536"
    if normalized in {"16:9", "21:9", "3:2", "4:3"}:
        return "1536x1024"
    return "1024x1024"


def load_inputs(workbook_path: Path) -> tuple[list[Product], list[HomeImage]]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if PRODUCT_SHEET not in workbook.sheetnames or HOME_SHEET not in workbook.sheetnames:
        raise KeyError(f"Required sheets missing from {workbook_path}")

    product_sheet = workbook[PRODUCT_SHEET]
    products: list[Product] = []
    for row in range(4, 79):
        title = str(product_sheet.cell(row, 3).value or "").strip()
        verdict = str(product_sheet.cell(row, 18).value or "").strip()
        if not title or verdict.casefold() == "écarter".casefold():
            continue
        prompt_text = str(product_sheet.cell(row, 21).value or "").strip()
        reference_dir = resolve_workbook_path(product_sheet.cell(row, 22).value)
        if not prompt_text:
            raise ValueError(f"Missing carousel prompts in row {row}: {title}")
        if not reference_dir.is_dir():
            raise FileNotFoundError(f"Reference folder missing in row {row}: {reference_dir}")
        split_product_prompts(prompt_text)
        products.append(
            Product(
                row=row,
                collection=str(product_sheet.cell(row, 2).value or "").strip(),
                title=title,
                verdict=verdict,
                prompt_text=prompt_text,
                reference_dir=reference_dir,
            )
        )

    home_sheet = workbook[HOME_SHEET]
    home_images: list[HomeImage] = []
    for row in range(5, home_sheet.max_row + 1):
        zone = str(home_sheet.cell(row, 1).value or "").strip()
        format_label = str(home_sheet.cell(row, 2).value or "").strip()
        prompt = str(home_sheet.cell(row, 3).value or "").strip()
        if zone and format_label and prompt:
            home_images.append(HomeImage(row=row, zone=zone, format_label=format_label, prompt=prompt))
    workbook.close()
    return products, home_images


def reference_images(reference_dir: Path) -> tuple[Path, ...]:
    paths = tuple(path for path in (reference_dir / "img_1.png", reference_dir / "img_2.png") if path.is_file())
    if not paths:
        raise FileNotFoundError(f"No img_1.png reference found in {reference_dir}")
    return paths


def build_jobs(
    products: list[Product],
    home_images: list[HomeImage],
    output_root: Path,
    limit: int,
    only_packshot: bool,
    include_home: bool,
    home_limit: int,
) -> list[Job]:
    selected_products = products[:limit] if limit > 0 else products
    jobs: list[Job] = []
    for product in selected_products:
        product_dir = output_root / f"{product.row:02d}_{slugify(product.title)}"
        for label, prompt in split_product_prompts(product.prompt_text):
            if only_packshot and label != "PACKSHOT":
                continue
            jobs.append(
                Job(
                    row_label=str(product.row),
                    product=product.title,
                    image_type=label,
                    prompt=f"{prompt}\n\n{LICENSE_GUARD}",
                    size="1024x1024",
                    output_path=product_dir / FILE_NAMES[label],
                    reference_paths=reference_images(product.reference_dir) if label in EDIT_LABELS else (),
                )
            )

    if include_home:
        selected_home = home_images[:home_limit] if home_limit > 0 else home_images
        home_dir = output_root / "_home_collections"
        for item in selected_home:
            jobs.append(
                Job(
                    row_label=f"home-{item.row}",
                    product=item.zone,
                    image_type="HOME/COLLECTION",
                    prompt=f"{item.prompt}\n\n{LICENSE_GUARD}",
                    size=format_to_size(item.format_label),
                    output_path=home_dir / f"{item.row:02d}_{slugify(item.zone)}.png",
                )
            )
    return jobs


def estimated_cost(quality: str, size: str) -> float:
    return COST_USD[quality][size]


def retry_call(call: Callable[[], object], max_retries: int) -> object:
    for attempt in range(max_retries + 1):
        try:
            return call()
        except Exception as exc:
            status = getattr(exc, "status_code", None)
            retryable = isinstance(exc, (RateLimitError, APIConnectionError, APITimeoutError)) or status in {
                429,
                500,
                502,
                503,
                504,
            }
            if not retryable or attempt >= max_retries:
                raise
            delay = min(60.0, (2**attempt) + random.uniform(0.25, 1.25))
            print(f"Retryable API error ({type(exc).__name__}); retry in {delay:.1f}s", file=sys.stderr)
            time.sleep(delay)
    raise RuntimeError("Retry loop exited unexpectedly")


def decode_response(response: object) -> bytes:
    data = getattr(response, "data", None)
    if not data or not getattr(data[0], "b64_json", None):
        raise ValueError("OpenAI image response did not contain b64_json")
    return base64.b64decode(data[0].b64_json)


def generate_job(client: OpenAI, job: Job, model: str, quality: str, max_retries: int) -> bytes:
    if job.reference_paths:
        def edit_call() -> object:
            with ExitStack() as stack:
                image_files = [stack.enter_context(path.open("rb")) for path in job.reference_paths]
                return client.images.edit(
                    model=model,
                    image=image_files,
                    prompt=job.prompt,
                    size=job.size,
                    quality=quality,
                    output_format="png",
                    input_fidelity="high",
                )

        return decode_response(retry_call(edit_call, max_retries))

    return decode_response(
        retry_call(
            lambda: client.images.generate(
                model=model,
                prompt=job.prompt,
                size=job.size,
                quality=quality,
                output_format="png",
            ),
            max_retries,
        )
    )


def validate_png(path: Path, expected_size: str) -> None:
    with Image.open(path) as image:
        image.verify()
    with Image.open(path) as image:
        expected = tuple(int(value) for value in expected_size.split("x"))
        if image.format != "PNG" or image.size != expected:
            raise ValueError(f"Invalid output {path}: format={image.format}, size={image.size}, expected={expected}")


def append_log(log_path: Path, rows: Iterable[dict[str, object]]) -> None:
    fieldnames = ["ligne", "produit", "type_image", "statut", "chemin", "coût_estimé"]
    log_path.parent.mkdir(parents=True, exist_ok=True)
    exists = log_path.exists()
    with log_path.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    products, home_images = load_inputs(args.workbook)
    jobs = build_jobs(
        products,
        home_images,
        args.output_root,
        args.limit,
        args.only_packshot,
        args.include_home,
        args.home_limit,
    )
    if not jobs:
        raise SystemExit("No image jobs selected")

    log_path = args.output_root / ("generation_log_dry_run.csv" if args.dry_run else "generation_log.csv")
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not args.dry_run:
        if not api_key:
            raise SystemExit(
                "OPENAI_API_KEY is not set in this process. Set it in the Codex environment, then rerun this command."
            )
        if api_key.startswith("export ") or "OPENAI_API_KEY=" in api_key or not api_key.startswith("sk-"):
            raise SystemExit(
                "OPENAI_API_KEY is malformed. It must contain only the secret value beginning with 'sk-', "
                "without 'export OPENAI_API_KEY=' or surrounding quotes."
            )
    client = None if args.dry_run else OpenAI(api_key=api_key)
    total_estimate = sum(estimated_cost(args.quality, job.size) for job in jobs)
    print(
        f"Plan: {len(jobs)} images, model={args.model}, quality={args.quality}, "
        f"estimated base cost=${total_estimate:.3f} (edit input-image tokens excluded)"
    )

    errors = 0
    for index, job in enumerate(jobs, start=1):
        cost = estimated_cost(args.quality, job.size)
        job.output_path.parent.mkdir(parents=True, exist_ok=True)
        status = "dry-run"
        authentication_failed = False
        try:
            if args.dry_run:
                pass
            elif job.output_path.exists() and not args.force:
                validate_png(job.output_path, job.size)
                status = "skipped-existing"
                cost = 0.0
            else:
                print(f"[{index}/{len(jobs)}] {job.row_label} {job.image_type}: {job.product}")
                image_bytes = generate_job(client, job, args.model, args.quality, args.max_retries)
                temp_path = job.output_path.with_suffix(".tmp.png")
                temp_path.write_bytes(image_bytes)
                validate_png(temp_path, job.size)
                temp_path.replace(job.output_path)
                status = "generated"
                if args.sleep > 0:
                    time.sleep(args.sleep)
        except AuthenticationError as exc:
            errors += 1
            authentication_failed = True
            status = f"error: {type(exc).__name__}: {exc}"
            cost = 0.0
            print(f"ERROR {job.row_label} {job.image_type}: API authentication failed.", file=sys.stderr)
        except Exception as exc:
            errors += 1
            status = f"error: {type(exc).__name__}: {exc}"
            cost = 0.0
            print(f"ERROR {job.row_label} {job.image_type}: {exc}", file=sys.stderr)

        append_log(
            log_path,
            [
                {
                    "ligne": job.row_label,
                    "produit": job.product,
                    "type_image": job.image_type,
                    "statut": status,
                    "chemin": str(job.output_path),
                    "coût_estimé": f"{cost:.3f}",
                }
            ],
        )
        if authentication_failed:
            print("Stopping after the first authentication error. Reload OPENAI_API_KEY and rerun.", file=sys.stderr)
            break

    print(f"Completed: {len(jobs) - errors}/{len(jobs)} jobs; log={log_path}")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
